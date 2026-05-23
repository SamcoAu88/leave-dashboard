import re
import openpyxl
import pandas as pd
from datetime import datetime, timedelta

LEAVE_TYPES = {
    "1":   "Annual Leave",
    "2":   "Long Service Leave",
    "3":   "48/52 Purchased Leave",
    "4":   "Other Leave",
    "M":   "Parental Leave",
    "T":   "Training",
    "TU":  "Time off in Lieu",
    "F":   "Annual Leave",
}

LEAVE_COLORS = {
    "Annual Leave":              "#378ADD",
    "Annual Leave (partial week)": "#85B7EB",
    "Long Service Leave":        "#1D9E75",
    "Parental Leave":            "#D4537E",
    "Time off in Lieu":          "#7F77DD",
    "Training":                  "#BA7517",
    "48/52 Purchased Leave":     "#D85A30",
    "Other Leave":               "#888780",
    "Unknown":                   "#B4B2A9",
}

QLD_PUBLIC_HOLIDAYS_2026 = [
    datetime(2026, 1, 1),  datetime(2026, 1, 26), datetime(2026, 4, 3),
    datetime(2026, 4, 6),  datetime(2026, 4, 25), datetime(2026, 5, 4),
    datetime(2026, 8, 12), datetime(2026, 12, 25), datetime(2026, 12, 28),
]

STAFF_START_ROW = 16
HEADER_ROW      = 15


def _find_first_date_col(ws):
    """Scan row 9 for the first real 2025/2026 datetime — that is col index of Mon Dec 29."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(9, c).value
        if isinstance(v, datetime) and v.year >= 2025:
            return c
    return 8   # fallback


def _normalise_vehicle(raw):
    if not raw:
        return "Unknown"
    s = str(raw).strip().lower()
    if "motor" in s or "bike" in s or "mbike" in s or "cycle" in s:
        return "Motorbike"
    if "edv" in s or "electric" in s:
        return "EDV"
    if "both" in s:
        return "Both"
    return str(raw).strip() or "Unknown"


# Partial-week day markers — these mean the person is on leave for PART of the week
# Format examples: "M-TU", "W-F", "TH-F", "M-TH", "T-F", "M-W", "TU-TH" etc.
# Days: M=Mon, T/TU=Tue, W=Wed, TH=Thu, F=Fri
_DAY_ORDER = {"M":1, "T":2, "TU":2, "W":3, "TH":4, "F":5}

def _parse_day_range(s):
    """Parse 'M-TH' style → (start_day, end_day) as ints 1-5. Returns None if not parseable."""
    s = re.sub(r"\s+", "", s.upper())
    # Single day codes
    single = {"M":1, "T":2, "TU":2, "W":3, "TH":4, "F":5}
    if s in single:
        return (single[s], single[s])
    # Range like M-TH, TU-F, W-F etc
    m = re.match(r"^([A-Z]+)-([A-Z]+)$", s)
    if m:
        d1, d2 = m.group(1), m.group(2)
        if d1 in single and d2 in single:
            return (single[d1], single[d2])
    return None

def _days_in_range(code):
    """Return number of days implied by a partial-week code."""
    result = _parse_day_range(code)
    if result:
        return max(1, result[1] - result[0] + 1)
    return 1

def _resolve_leave_type(val):
    """Map raw cell value to a leave type label."""
    if val is None or val == 0 or val == "":
        return None
    s = str(val).strip()
    upper = s.upper().replace(" ", "")

    # Numeric codes: 1=Annual, 2=LSL, 3=48/52, 4=Other
    # Higher numbers (5-16) = day counts entered as numbers → treat as Annual Leave days
    num_map = {
        "1":"Annual Leave", "2":"Long Service Leave",
        "3":"48/52 Purchased Leave", "4":"Other Leave",
    }
    if upper in num_map:
        return num_map[upper]
    # Numbers 5-20: likely day counts → Annual Leave
    try:
        n = int(upper)
        if 5 <= n <= 20:
            return "Annual Leave"
    except ValueError:
        pass

    # Single letter codes
    single_map = {"M":"Parental Leave", "T":"Training", "TU":"Time off in Lieu"}
    if upper in single_map:
        return single_map[upper]

    # Partial-week day ranges: M-TH, W-F, TH-F, T-F, M-TU, M-W, TUE-F etc.
    if _parse_day_range(upper) is not None:
        return "Annual Leave (partial week)"

    # Fallback
    return "Annual Leave (partial week)"


def _col_to_week_start(col_idx, first_date_col):
    """
    Layout: first_date_col = Monday of week 0 (Dec 29 2025).
    Every 6 columns = 1 week:
      offset 0 → week marker (same Monday repeated)
      offset 1 → Mon, 2 → Tue … 5 → Fri
    """
    start_date = datetime(2025, 12, 29)
    rel        = col_idx - first_date_col
    week_num   = rel // 6
    day_offset = rel % 6
    week_start = start_date + timedelta(weeks=week_num)
    return week_start, day_offset


def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Sheet1"]

    first_date_col = _find_first_date_col(ws)

    staff_list = []
    for row_idx in range(STAFF_START_ROW, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        if not name or str(name).strip() == "":
            continue                          # skip blank rows, don't break

        depot        = str(ws.cell(row_idx, 2).value or "").strip()
        aps          = str(ws.cell(row_idx, 3).value or "").strip()
        designation  = str(ws.cell(row_idx, 4).value or "").strip()
        route_team   = str(ws.cell(row_idx, 5).value or "").strip()
        vehicle_type = _normalise_vehicle(ws.cell(row_idx, 6).value)
        pt_hours     = ws.cell(row_idx, 7).value

        leave_entries = []
        seen_weeks    = set()

        for col_idx in range(first_date_col, ws.max_column + 1):
            val = ws.cell(row_idx, col_idx).value
            if val is None or val == 0 or val == "":
                continue

            week_start, day_offset = _col_to_week_start(col_idx, first_date_col)
            wk_key = week_start.strftime("%Y-%m-%d")

            leave_type = _resolve_leave_type(val)
            if not leave_type:
                continue

            if day_offset == 0:
                # week-marker column → full week leave
                if wk_key in seen_weeks:
                    continue
                seen_weeks.add(wk_key)
                actual_date = week_start
                is_week     = True
            else:
                actual_date = week_start + timedelta(days=day_offset - 1)
                is_week     = str(val).strip() == "1"

            leave_entries.append({
                "week_start": week_start,
                "date":       actual_date,
                "day_offset": day_offset,
                "raw_value":  str(val),
                "leave_type": leave_type,
                "is_week":    is_week,
            })

        staff_list.append({
            "name":         str(name).strip(),
            "depot":        depot,
            "aps":          aps,
            "designation":  designation,
            "route_team":   route_team,
            "vehicle_type": vehicle_type,
            "pt_hours":     pt_hours,
            "leaves":       leave_entries,
        })

    return staff_list


def build_leave_df(staff_list):
    rows = []
    for staff in staff_list:
        for entry in staff["leaves"]:
            wk        = entry["week_start"]
            iso_week  = f"W{wk.isocalendar()[1]:02d} {wk.year}"
            rows.append({
                "name":         staff["name"],
                "depot":        staff["depot"],
                "aps":          staff["aps"],
                "designation":  staff["designation"],
                "route_team":   staff["route_team"],
                "vehicle_type": staff["vehicle_type"],
                "week_start":   wk,
                "iso_week":     iso_week,
                "leave_type":   entry["leave_type"],
                "is_week":      entry["is_week"],
                "month":        wk.strftime("%b %Y"),
                "month_num":    wk.month,
                "year":         wk.year,
            })
    return pd.DataFrame(rows)


def get_all_weeks():
    weeks, d = [], datetime(2025, 12, 29)
    end = datetime(2026, 12, 28)
    while d <= end:
        weeks.append(d)
        d += timedelta(weeks=1)
    return weeks


def get_week_labels(weeks):
    return [f"W{w.isocalendar()[1]:02d} {w.year}" for w in weeks]


def concurrent_by_week(df, weeks):
    rows = []
    for wk in weeks:
        iso_wk        = f"W{wk.isocalendar()[1]:02d} {wk.year}"
        staff_on_leave = df[df["week_start"] == wk]["name"].unique().tolist()
        rows.append({
            "week_start":       wk,
            "iso_week":         iso_wk,
            "concurrent_count": len(staff_on_leave),
            "staff_on_leave":   staff_on_leave,
            "month":            wk.strftime("%b %Y"),
        })
    return pd.DataFrame(rows)
